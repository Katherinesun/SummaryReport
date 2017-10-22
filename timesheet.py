import sys
from collections import defaultdict
from openpyxl import load_workbook, Workbook

delimiter = ','

# Output file header
TIMESHEET_HEADER = ['Employee',
                    'Name',
                    'Cost Center',
                    'Mileint',
                    'Mileout',
                    'KM>10',
                    'Ord',
                    'Training',
                    'Mkup',
                    'Total',
                    'Sat Loading',
                    'Sun Loading',
                    'SatCasual Loading',
                    'SunCasual Loading',
                    'ES',
                    'NS',
                    'PH',
                    'PH Loading',
                    'PHNW',
                    'AL',
                    'LL',
                    'PCL',
                    'Compassionate Leave',
                    'OT1x5',
                    'OT2',
                    'OT2x5',
                    'Internet']

PAYSHT_NAME = [None,
               None,
               None,
               'MILEINT',
               'MILEOUT',
               'KM>10',
               'ORD',
               'TRAINING',
               'Mkup',
               'Total',
               'Sat Loading',
               'Sun Loading',
               'SATCAS',
               'SUNCAS',
               'ES',
               'NS',
               'PH',
               'PHLOAD',  # 'PHCAS' will be treated the same as 'PHLOAD'
               'PHNW',
               'AL',
               'LL',
               'PCL',
               'Compassionate Leave',
               'OT1x5',
               'OT2',
               'OT2x5',
               'INTERNET']


TIMESHEET_PAYSHT_MAP = dict(zip(TIMESHEET_HEADER, PAYSHT_NAME))

SUMMARY_HEADER = list(TIMESHEET_HEADER)
SUMMARY_HEADER.remove('Cost Center')
SUMMARY_HEADER.remove('KM>10')
SUMMARY_HEADER.remove('Total')


def to_float(var):
    try:
        return float(var)
    except ValueError:
        return float(0)

def get_employee_list(filename):
    """
    Read the DCW_file to get the employee information, and return a dictionary
    with mapping of employee code and employee full name.
    - Assuming the first column is Employee Code
    - Assuming the second column is Full Name
    @TODO: Check the header to work out which columns to Check

    @var filename: the path of the DCW file (Excel spread sheet)
    @return employee_dict: a dictionary with {employee_code: employee_name}
    """
    employees = {}
    wb = load_workbook(filename, read_only=True)
    ws = wb.active
    # Assuming the first column is Employee Code
    # Assuming the second column is Full Name
    # @TODO: Check the header to work out which columns to Check
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if isinstance(cell.value, long):
                employees[str(cell.value)] = ws.cell(row=cell.row, column=2).value
    return employees


class CostCenter:
    """
    Put all the data related to a cost center into a class.
    An instance of CostCenter belongs to an instance of IndividualEmployeeTimeSheet
    """
    PAYTYPES = filter(bool, PAYSHT_NAME)

    def __init__(self, employee_timesheet, cost_center):
        self.employee_timesheet = employee_timesheet
        self.cost_center = cost_center
        self.data = dict(zip(self.PAYTYPES, [''] * len(self.PAYTYPES)))

    def calculate_total(self):
        """
        Calculate the total cost for a cost center for an employee:
          'KM>10' + 'Ord' + 'TRAINING'
        """
        self.data['Total'] = (to_float(self.data['KM>10']) +
                              to_float(self.data['ORD']) +
                              to_float(self.data['TRAINING']))

    def process(self, cols):
        """
        Process columns of a row (for this cost center) from PAYSHT.INP
        """
        paytype = cols[5]
        value = float(cols[6])

        if paytype == 'PHCAS':
            # 'PHCAS' is treated the same as 'PHLOAD'
            paytype = 'PHLOAD'

        if (paytype == 'ORD') and (value < 1) and (value != 0.5):
            self.data['KM>10'] = value
            self.calculate_total()
        elif paytype in self.data:
            self.data[paytype] = value
            self.calculate_total()
        else:
            print("Paytype '%s' is not recognized." % paytype,
                  "This row is not processed: %s" % ','.join(cols))

    def __str__(self):
        text = self.cost_center
        for k in TIMESHEET_HEADER:
            v = TIMESHEET_PAYSHT_MAP[k]
            if v in self.data:
                text += delimiter + str(self.data[v])
        return text

    def get_timesheet(self):
        timesheet = [self.cost_center]
        for k in TIMESHEET_HEADER:
            v = TIMESHEET_PAYSHT_MAP[k]
            if v in self.data:
                timesheet.append(self.data[v])
        return timesheet


class IndividualEmployeeTimeSheet:
    def __init__(self, employee_code, employee_name):
        self.employee_code = employee_code
        self.employee_name = employee_name
        self.cost_center_list = {}

    def process(self, cols):
        """
        Process columns of a row from PAYSHT.INP
        @var cols: a list of columns of a row/line
        """
        cost_center_id = cols[12]
        if cost_center_id not in self.cost_center_list:
            self.cost_center_list[cost_center_id] = CostCenter(self,
                                                               cost_center_id)
        self.cost_center_list[cost_center_id].process(cols)

    def __str__(self):
        text = ""
        first_row = True
        if delimiter == ',':
            employee_name = self.employee_name.replace(',', '')
        for cc in self.cost_center_list.values():
            if first_row:
                text += delimiter.join([str(self.employee_code),
                                        str(employee_name),
                                        str(cc)])
                first_row = False
            else:
                text += delimiter.join(['', '', str(cc)])
            text += '\n'
        return text

    def get_timesheet(self):
        timesheet = []
        first_row = True
        for cc in self.cost_center_list.values():
            if first_row:
                timesheet.append([self.employee_code, self.employee_name] +
                                  cc.get_timesheet())
                first_row = False
            else:
                timesheet.append(['', ''] + cc.get_timesheet())
        return timesheet

    def calculate_summary(self):
        self.summary = dict(zip(SUMMARY_HEADER, [''] * len(SUMMARY_HEADER)))
        self.summary['Employee'] = self.employee_code
        self.summary['Name'] = self.employee_name
        cost_centers = self.cost_center_list.values()
        for k in self.summary:
            v = TIMESHEET_PAYSHT_MAP[k]
            if v in CostCenter.PAYTYPES:
                total = sum(map(to_float, [cc.data[v] for cc in cost_centers]))
                if total == 0:
                    self.summary[k] = ''
                else:
                    self.summary[k] = total
        return self.summary

    def get_summary(self):
        self.calculate_summary()
        return [self.summary[k] for k in SUMMARY_HEADER]

    def get_summary_text(self):
        summary = dict(self.calculate_summary())
        if delimiter == ',':
            summary['Name'] = self.employee_name.replace(',', '')
        text = delimiter.join(map(lambda k: str(summary[k]),
                                  SUMMARY_HEADER))
        text += '\n'
        return text


class TimeSheet:

    def __init__(self, employee_list):
        self.employee_list = employee_list
        self.all_employee_timesheets = {}

    def __setitem__(self, key, value):
        self.all_employee_timesheets[key] = value

    def __getitem__(self):
        return self.all_employee_timesheets[key]

    def load_from_paysht_inp(self, input_filename):
        self.input_filename = input_filename
        with open(input_filename, 'r') as fsock:
            for line in fsock:
                cols = line.split(',')
                employee_code = cols[2]
                if employee_code not in self.all_employee_timesheets:
                    if employee_code in self.employee_list:
                        employee_name = self.employee_list[employee_code]
                    else:
                        print ("Unable to find the name for employee %s" %
                               employee_code)
                        employee_name = ""
                    ind_timesheet = IndividualEmployeeTimeSheet(
                                    employee_code,
                                    employee_name)
                    self.all_employee_timesheets[employee_code] = ind_timesheet
                self.all_employee_timesheets[employee_code].process(cols)


    def save_to_text_file(self, output_filename):
        with open(output_filename, 'w+') as fsock:
            # Output timesheet
            fsock.write("Timesheet\n" + "Pay Period:\n")
            fsock.write(delimiter.join(TIMESHEET_HEADER) + '\n')
            for employee in sorted(self.all_employee_timesheets.keys()):
                ts = self.all_employee_timesheets[employee]
                fsock.write(str(ts))
            fsock.write("\n\n\n")
            fsock.flush()

            # output summary report
            fsock.write("Summary Report\n" + "Pay Period:\n")
            fsock.write(delimiter.join(SUMMARY_HEADER) + '\n')
            for employee in sorted(self.all_employee_timesheets.keys()):
                ts = self.all_employee_timesheets[employee]
                summary = ts.get_summary_text()
                fsock.write(str(summary))
            fsock.flush()


    def save_to_excel(self, output_filename):
        if not output_filename.endswith('.xlsx'):
            print "Must save the excel output to a file with .xlsx extension"
            sys.exit(1)
        wb = Workbook()
        ws = wb.active

        # Write the title and header for Timesheet
        ws.append(['Timesheet'])
        ws.append(TIMESHEET_HEADER)

        # Write the contents for Timesheet
        for employee in sorted(self.all_employee_timesheets.keys()):
            ts = self.all_employee_timesheets[employee]
            for row in ts.get_timesheet():
                ws.append(row)

        # Write 3 blank rows
        for _ in range(3):
            ws.append([])

        # Write the title and header for summary report
        ws.append(['Summary Report'])
        ws.append(['Pay Period:'])
        ws.append(SUMMARY_HEADER)

        # Write the contents for summary report
        for employee in sorted(self.all_employee_timesheets.keys()):
            ts = self.all_employee_timesheets[employee]
            summary = ts.get_summary()
            ws.append(summary)

        # Save to the file
        wb.save(output_filename)
